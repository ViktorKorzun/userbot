from telethon import TelegramClient, events
import os
import re
from openpyxl import Workbook, load_workbook

# === Налаштування ===
api_id = 24063327
api_hash = "4b78510032991cc80972c8e91235df51"
chat_username = "https://t.me/orendakvarturlviv"  # назва групи або юзернейм
file_name = "phones.xlsx"

# === Ініціалізація клієнта ===
client = TelegramClient("userbot", api_id, api_hash)

# Регулярний вираз для пошуку українських номерів
# +380XXXXXXXXX, 380XXXXXXXXX або 0XXXXXXXXX
phone_pattern = re.compile(r"(?:\+?380|0)\d{9}")

def init_excel():
    """Створення Excel-файлу з заголовками, якщо він не існує"""
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Phones"
        ws.append(["ID", "Відправник", "Телефон"])
        wb.save(file_name)


def phone_exists(msg_id):
    """Перевірка, чи є повідомлення у файлі"""
    wb = load_workbook(file_name)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == msg_id:
            return True
    return False


def save_phone_to_excel(msg_id, sender_id, phone):
    """Збереження номера в Excel"""
    if not phone_exists(msg_id):
        wb = load_workbook(file_name)
        ws = wb.active
        ws.append([msg_id, sender_id, phone])
        wb.save(file_name)


async def export_old_messages(entity):
    """Експорт старих повідомлень із номерами"""
    print("Пошук старих українських номерів...")
    count = 0
    async for message in client.iter_messages(entity, limit=None):
        if message.message:
            phones = phone_pattern.findall(message.message)
            for phone in phones:
                save_phone_to_excel(message.id, message.sender_id, phone)
                count += 1
    print(f"✅ Знайдено {count} номерів.")


@client.on(events.NewMessage)
async def new_message_handler(event):
    """Обробка нових повідомлень із номерами"""
    if event.chat and event.chat.username == chat_username:
        text = event.message.message
        if text:
            phones = phone_pattern.findall(text)
            for phone in phones:
                save_phone_to_excel(event.message.id, event.sender_id, phone)
                print(f"[НОВИЙ НОМЕР] {event.sender_id}: {phone}")


async def main():
    init_excel()
    await client.start()
    print("Юзербот запущений...")

    entity = await client.get_entity(chat_username)
    print(f"✅ Підключено до чату: {entity.title} (ID: {entity.id})")

    await export_old_messages(entity)
    print("Очікування нових повідомлень...")
    await client.run_until_disconnected()


client.loop.run_until_complete(main())
